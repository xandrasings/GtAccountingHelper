from .constants import *
from .dateTime import *
from .fileManagement import *
from ..Classes.AmazonNonOrderRecord import *
from ..Classes.AmazonOrderRecord import *
from ..Classes.CutOffSplit import *
from ..Classes.QuickBooksRecord import *

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import itertools


def loadWorkbook(fileType, filePath, readOnly = False):
	try:
		return load_workbook(filePath, readOnly)
	except:
		quit('Could not open workbook for ' + fileType + '.')


def loadSheet(workbook, fileType):
	sheet = None
	if SHEET not in LOCATION[fileType] or len(workbook.sheetnames) == 1:
		try:
			sheet = workbook.get_active_sheet()
		except:
			quit('Could not open active sheet of ' + fileType + ' workbook.')
	else:
		sheetName = LOCATION[fileType][SHEET]
		try:
			sheet = workbook[sheetName]
		except:
			quit('Could not open sheet ' + sheetName + ' of ' + fileType + ' workbook.')

	verifySheetFormatting(fileType, sheet)
	return sheet

def verifySheetFormatting(fileType, sheet):
	verifyColumnMax(fileType, sheet)
	verifyHeaderContent(fileType, sheet)

def verifyColumnMax(fileType, sheet):
	actualColumnMax = sheet.max_column
	expectedColumnMax = LOCATION[fileType][COLUMN][MAX]
	if actualColumnMax != expectedColumnMax:
		quit(fileType + ' workbook had ' + str(actualColumnMax) + ' column max instead of expected value ' + str(expectedColumnMax) + '.')

def verifyHeaderContent(fileType, sheet):
	row = LOCATION[fileType][ROW][HEADER]
	for column, expectedContent in LOCATION[fileType][HEADER].items():
		if getCellString(sheet, row, column) != expectedContent.upper():
			quit(fileType + ' workbook did not have expected column header ' + expectedContent + ' in row ' + str(row) + ' column ' + str(column) + '.')


def getCellValue(sheet, row, column):
	try:
		return sheet.cell(row = row, column = column).value
	except:
		quit('Could not get cell value for row ' + str(row) + ', column ' + str(column) + '.')


def getCellString(sheet, row, column):
	cellValue = getCellValue(sheet, row, column)
	try:
		return str(cellValue).upper().strip()
	except:
		quit('Could not convert cell value for row ' + str(row) + ', column ' + str(column) + ' to string.')


def getCellFloat(sheet, row, column):
	cellValue = getCellValue(sheet, row, column)
	try:
		if cellValue == None:
			cellValue = 0
		return float(cellValue)
	except:
		quit('Could not convert cell value ' + cellValue + ' to float.')


def getCellDateString(fileType, sheet, row, column):
	dateString = getCellString(sheet, row, column)
	try:
		return datetime.strptime(dateString[0 : dateString.index(',') + 6], '%b %d, %Y')
	except:
		quit('Could not convert date string ' + dateString + ' to datetime.')


def setCellValue(sheet, row, column, value):
	sheet.cell(row = row, column = column).value = value;


def setCellColor(sheet, row, column, color):
	fill = PatternFill(SOLID, fgColor = color)
	sheet.cell(row = row, column = column).fill = fill

def setRowColor(fileType, sheet, row, color):
	for column in range (LOCATION[fileType][COLUMN][MIN], LOCATION[fileType][COLUMN][NEW_MAX] + 1):
		setCellColor(sheet, row, column, color)

def processQuickBooksReport(filePath):
	workbook = loadWorkbook(QUICKBOOKS, filePath, True)
	sheet = loadSheet(workbook, QUICKBOOKS)

	invoices = []

	for row in range(LOCATION[QUICKBOOKS][ROW][HEADER] + 2, sheet.max_row + 1): # TODO 
		date = getCellValue(sheet, row, LOCATION[QUICKBOOKS][COLUMN][DATE])
		invoiceNumber = getCellString(sheet, row, LOCATION[QUICKBOOKS][COLUMN][INVOICE_NUMBER])
		city = getCellString(sheet, row, LOCATION[QUICKBOOKS][COLUMN][CITY])
		debit = getCellFloat(sheet, row, LOCATION[QUICKBOOKS][COLUMN][DEBIT])
		credit = getCellFloat(sheet, row, LOCATION[QUICKBOOKS][COLUMN][CREDIT])

		invoices.append(QuickBooksRecord(date, invoiceNumber, city, debit, credit))

	return invoices


def processAmazonReport(amazonFilePath, quickBooksRecords, unavailableBalance, balanceAction, reportFilePath):
	workbook = loadWorkbook(AMAZON, amazonFilePath, False)
	sheet = loadSheet(workbook, AMAZON)

	orders = []
	nonOrders = []

	for row in range(LOCATION[AMAZON][ROW][HEADER] + 1, sheet.max_row + 1):
		recordType = getCellString(sheet, row, LOCATION[AMAZON][COLUMN][TYPE]) 
		date = getCellDateString(AMAZON, sheet, row, LOCATION[AMAZON][COLUMN][DATE])

		if recordType == ORDER:
			city = getCellString(sheet, row, LOCATION[AMAZON][COLUMN][CITY])
			orderId = getCellString(sheet, row, LOCATION[AMAZON][COLUMN][ORDER_ID])
			productSales = getCellFloat(sheet, row, LOCATION[AMAZON][COLUMN][PRODUCT_SALES])
			shippingCredits = getCellFloat(sheet, row, LOCATION[AMAZON][COLUMN][SHIPPING_CREDITS])
			productSalesTaxCollected = getCellFloat(sheet, row, LOCATION[AMAZON][COLUMN][PRODUCT_SALES_TAX])
			shippingCreditsTaxCollected = getCellFloat(sheet, row, LOCATION[AMAZON][COLUMN][SHIPPING_CREDITS_TAX])
			giftwrapCreditsTaxCollected = getCellFloat(sheet, row, LOCATION[AMAZON][COLUMN][GIFT_WRAP_CREDITS_TAX])
			promotionalRebatesTaxCollected = getCellFloat(sheet, row, LOCATION[AMAZON][COLUMN][PROMOTIONAL_REBATES_TAX])
			salesTaxCollected = productSalesTaxCollected + shippingCreditsTaxCollected + giftwrapCreditsTaxCollected + promotionalRebatesTaxCollected
			sellingFees = getCellFloat(sheet, row, LOCATION[AMAZON][COLUMN][SELLING_FEES])
			total = getCellFloat(sheet, row, LOCATION[AMAZON][COLUMN][TOTAL])
			orders.append(AmazonOrderRecord(row, date, orderId, city, productSales, shippingCredits, salesTaxCollected, sellingFees, total))
		else:
			nonOrders.append(AmazonNonOrderRecord(row, date, recordType))
	
	populateInvoiceNumbers(orders, quickBooksRecords)
	orders.sort(reverse = True)
	cutOffSplit = identifyCutOffRecords(orders, unavailableBalance) if balanceAction else identifyEstimatedCutOffRecords(orders, unavailableBalance)

	modifyAmazonReport(sheet, orders, nonOrders, cutOffSplit)
	saveAmazonReport(workbook, reportFilePath)
	workbook.save(reportFilePath)


def populateInvoiceNumbers(amazonOrderRecords, quickBooksRecords):
	orderBreakDown = {}
	for amazonOrder in amazonOrderRecords:
		orderId = amazonOrder.getOrderId()
		if orderId not in orderBreakDown:
			orderBreakDown[orderId] = []
		orderBreakDown[orderId].append(amazonOrder)

	amazonDetailedBreakdown = {}
	for orderId, amazonOrderRecords in orderBreakDown.items():
		city = amazonOrderRecords[0].getCity()
		matchSubtotal = round(sum(amazonOrder.getMatchSubtotal() for amazonOrder in amazonOrderRecords), 2)
		if city not in amazonDetailedBreakdown:
			amazonDetailedBreakdown[city] = {}
		if matchSubtotal not in amazonDetailedBreakdown[city]:
			amazonDetailedBreakdown[city][matchSubtotal] = []
		amazonDetailedBreakdown[city][matchSubtotal].append(amazonOrderRecords)

	quickBooksDetailedBreakdown = {}
	for quickBooksRecord in quickBooksRecords:
		city = quickBooksRecord.getCity()
		total = quickBooksRecord.getTotal()
		if city not in quickBooksDetailedBreakdown:
			quickBooksDetailedBreakdown[city] = {}
		if total not in quickBooksDetailedBreakdown[city]:
			quickBooksDetailedBreakdown[city][total] = []
		quickBooksDetailedBreakdown[city][total].append(quickBooksRecord)

	for city, totalBreakdown in amazonDetailedBreakdown.items():
		for total, amazonOrderGroups in totalBreakdown.items():
			try:
				quickBooksOrders = quickBooksDetailedBreakdown[city][total]
				matchCount = len(amazonOrderGroups)
				pairings = []
				for quickBooksMatches in itertools.permutations(quickBooksOrders, matchCount):
					pairings.append(quickBooksMatches)

				bestEditDistance = 20 * matchCount
				for quickBookMatches in pairings:
					newEditDistance = 0
					for i in range(0, matchCount):
						newEditDistance = newEditDistance + calculateMatchEditDistance(quickBookMatches[i], amazonOrderGroups[i])
					if newEditDistance < bestEditDistance:
						bestEditDistance = newEditDistance
						for i in range(0, matchCount):
							for amazonOrder in amazonOrderGroups[i]:
								amazonOrder.setInvoiceNumber(quickBookMatches[i].getInvoiceNumber())
			except KeyError:
				output('Key error for city ' + city + ' and total ' + str(total))
				break
			except:
				break


def calculateMatchEditDistance(quickBooksMatch, amazonMatches):
	return calculateDateEditDistance(quickBooksMatch.getDate(), amazonMatches[0].getDate())


def calculateDateEditDistance(date1, date2):
	subtractDay(date1, date2)
	return abs(subtractDay(date1, date2))


def identifyCutOffRecords(orders, unavailableBalance):
	if unavailableBalance > 0:
		cutOffValues = attemptCutOff(orders, unavailableBalance)

		if len(cutOffValues) == 0:
			return identifyEstimatedCutOffRecords(orders, unavailableBalance)
		else:
			identifyExactCutOffRecords(orders, cutOffValues)
			return None


def identifyExactCutOffRecords(orders, cutOffValues):
	for value in cutOffValues:
		for order in orders:
			if order.getTaxedTotal() == value and not order.isCutOff():
				order.setCutOff()
				break


def identifyEstimatedCutOffRecords(orders, unavailableBalance):
	for order in orders:
		taxedTotal = order.getTaxedTotal()
		order.setCutOff()
		if taxedTotal < unavailableBalance:
			unavailableBalance = unavailableBalance - taxedTotal
		else: 
			return CutOffSplit(order.getRow(), order.getTaxedTotal(), round(unavailableBalance,2))
	return None


def attemptCutOff(orders, unavailableBalance):
	pot = []
	for order in orders:
		pot.append(order.getTaxedTotal())
		goal = round(sum(pot) - unavailableBalance, 2)
		
		if goal == 0:
			return pot
		elif goal > 0:
			dropCountLimit = min(len(pot), DROP_COUNT_LIMIT)
			for dropCount in range(1, dropCountLimit):
				for drops in itertools.combinations(pot, dropCount):
					if round(sum(drops), 2) == goal:
						drops = list(drops)
						for drop in drops:
							pot.remove(drop)
						return pot
	return []


def saveAmazonReport(workbook, filePath):
	workbook.save(filePath)
	clearScreen()
	output('Report saved to ' + filePath)


def modifyAmazonReport(sheet, orders, nonOrders, cutOffSplit):
	markCutOffRecords(sheet, orders)
	markCutOffSplit(sheet, cutOffSplit)
	addNewDataColumns(sheet, orders)
	copyNonOrders(sheet, nonOrders)
	removeNonOrders(sheet, nonOrders)
	removeHeaderRows(sheet)


def markCutOffRecords(sheet, orders):
	for order in orders:
		if order.isCutOff():
			setRowColor(AMAZON, sheet, order.getRow(), CUT_OFF_COLOR)
	pass # TADA color the records where cutOff = true #MARK_CUT_OFF_RECORDS


def markCutOffSplit(sheet, cutOffSplit):
	if cutOffSplit != None:
		row = cutOffSplit.getRow()
		column = LOCATION[AMAZON][COLUMN][CUT_OFF_SPLIT]
		value = '* partial unavailable balance of $' + str(cutOffSplit.getUnavailableBalance()) + ' out of $' + str(cutOffSplit.getTaxedTotal()) + '*' 
		
		setCellValue(sheet, row, column, value)
		setCellColor(sheet, row, column, CUT_OFF_COLOR)


def addNewDataColumns(sheet, orders):
	addHeaders(sheet)
	addNewData(sheet, orders)


def addHeaders(sheet):
	for header in [INVOICE_NUMBER, CASH_RECEIVED]:
		addHeader(sheet, header)


def addHeader(sheet, header):
	setCellValue(sheet, LOCATION[AMAZON][ROW][HEADER], LOCATION[AMAZON][COLUMN][header], header)


def addNewData(sheet, orders):
	for order in orders:
		setCellValue(sheet, order.getRow(), LOCATION[AMAZON][COLUMN][CASH_RECEIVED], order.getCashReceived())
		setCellValue(sheet, order.getRow(), LOCATION[AMAZON][COLUMN][INVOICE_NUMBER], order.getInvoiceNumber())


def copyNonOrders(sheet, nonOrders):
	row = sheet.max_row + LOCATION[AMAZON][ROW][NON_ORDER_BUFFER]
	setCellValue(sheet, row, LOCATION[AMAZON][COLUMN][MIN], 'Non-Order Records:')
	for nonOrder in sorted(nonOrders):
		row = row + 1
		for column in range(LOCATION[AMAZON][COLUMN][MIN], LOCATION[AMAZON][COLUMN][MAX] + 1):
			setCellValue(sheet, row, column, getCellValue(sheet, nonOrder.getRow(), column))
		

def removeNonOrders(sheet, nonOrders):
	for nonOrder in sorted(nonOrders, key = lambda x: x.row, reverse = True):
		sheet.delete_rows(nonOrder.row)


def removeHeaderRows(sheet):
	sheet.delete_rows(1, 7)

